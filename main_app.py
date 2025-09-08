import re
import os
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import PyPDF2
import eel # Import Eel library

# Try to import PyMuPDF with error handling for package conflicts
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except (ImportError, RuntimeError) as e:
    print(f"PyMuPDF import failed: {e}")
    print("Falling back to PyPDF2 only")
    PYMUPDF_AVAILABLE = False

# Initialize Eel with the web folder (where your HTML, CSS, JS will be)
# Make sure to create a folder named 'web' in the same directory as this Python script
eel.init('web')

@eel.expose # Expose this function to JavaScript
def extract_text_from_pdf_eel(pdf_data_base64, method='auto'):
    """
    Extract text from PDF file (received as base64 string) using different methods.
    Returns the extracted text.
    """
    text = ""
    try:
        # Decode base64 PDF data to bytes
        pdf_bytes = base64.b64decode(pdf_data_base64)
        pdf_file = io.BytesIO(pdf_bytes) # Create a file-like object from bytes

        # Auto-select method based on availability
        if method == 'auto':
            method = 'pymupdf' if PYMUPDF_AVAILABLE else 'pypdf2'
        
        if method == 'pymupdf' and PYMUPDF_AVAILABLE:
            doc = fitz.open(stream=pdf_file.read(), filetype="pdf") # Use stream for PyMuPDF
            for page_num in range(doc.page_count):
                page = doc[page_num]
                text += page.get_text()
            doc.close()
            print(f"Extracted text using PyMuPDF from {doc.page_count} pages")
            
        elif method == 'pypdf2' or not PYMUPDF_AVAILABLE:
            pdf_reader = PyPDF2.PdfReader(pdf_file) # PyPDF2 can directly use file-like object
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text += page.extract_text()
            print(f"Extracted text using PyPDF2 from {len(pdf_reader.pages)} pages")
            
    except Exception as e:
        print(f"Error extracting text with {method}: {str(e)}")
        if method == 'pymupdf' and PYMUPDF_AVAILABLE:
            print("Trying PyPDF2 as fallback in Eel...")
            return extract_text_from_pdf_eel(pdf_data_base64, 'pypdf2')
        else:
            raise Exception(f"Failed to extract text from PDF: {str(e)}")
    
    return text

@eel.expose # Expose this function to JavaScript
def parse_pdf_data_to_excel_eel(pdf_data_base64, original_filename=""):
    """
    Reads PDF data (as base64), extracts text, and creates Excel with merged cells.
    Returns the Excel file as a base64 string.
    """
    print(f"Processing PDF data for filename: {original_filename}")
    
    try:
        pdf_text = extract_text_from_pdf_eel(pdf_data_base64, 'auto')
    except Exception as e:
        print(f"Error extracting text in Eel: {e}")
        return {"error": str(e)} # Return error to JS

    if not pdf_text.strip():
        return {"error": "No text could be extracted from the PDF data."}
    
    print(f"Extracted text length: {len(pdf_text)} characters")
    print("First 500 characters of extracted text:")
    print("-" * 50)
    print(pdf_text[:500])
    print("-" * 50)
    
    # Call the existing parsing function (now local to this script)
    # This function is not exposed to Eel directly, it's called internally
    excel_buffer = parse_text_to_excel_with_merging(pdf_text, original_filename)
    
    # Encode the Excel buffer to base64 for transfer to JavaScript
    excel_base64 = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
    print("Excel file generated and encoded to base64.")
    return {"success": True, "excel_data_base64": excel_base64, "filename": original_filename}


def parse_text_to_excel_with_merging(pdf_text, source_file=""):
    """
    Parses PDF text and creates Excel with merged cells for better readability.
    Returns an in-memory BytesIO object containing the Excel file.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "App Store Features"

    # Define the headers
    headers = ["Level 1 Heading", "Level 2 Heading", "Level 3 Heading", "List Item No.", "List Item Description"]
    sheet.append(headers)

    # Style the headers
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    # Improved regex patterns with more flexibility
    pattern_level1 = re.compile(r"^\s*(\d+(?:\s+.+)?)$", re.MULTILINE)
    pattern_level2 = re.compile(r"^\s*(\d+\.\d+(?:\s+.+)?)$", re.MULTILINE)
    pattern_level3 = re.compile(r"^\s*(\d+\.\d+\.\d+(?:\s+.+)?)$", re.MULTILINE)
    pattern_list_item = re.compile(r"^\s*(\d+)\.\s+(.+)$", re.MULTILINE)
    # Robust Regex for bullet points. Includes common unicode bullets and symbols.
    # Using `^[\s\u00A0]*` to match leading whitespace including non-breaking space
    # \u2022: Bullet (•), \u2023: Triangular Bullet (‣), \u25CF: Black Circle (●), \uF0B7: Wingdings bullet (), *: Asterisk, -: Hyphen, \u00B7: Middle Dot (·)
    pattern_bullet_item = re.compile(r"^[ \t]*[\u2022\u2023\u25CF\uF0B7\*\-\u00B7]\s*(.+)$", re.MULTILINE)


    parsed_data = []
    current_level1 = ""
    current_level2 = ""
    current_level3 = ""
    current_list_item_buffer = None # This buffer will now hold details for either numbered or bulleted items

    lines = pdf_text.split('\n')
    print(f"Processing {len(lines)} lines...")

    for line_num, raw_line in enumerate(lines, 1): # Changed to raw_line for debugging
        line = raw_line.strip()
        
        # Console logging for debugging (similar to Python)
        # print(f"\n--- Line {line_num} ---")
        # print(f"Raw: {repr(raw_line)}") # Use repr() to show invisible characters
        # print(f"Stripped: {repr(line)}")
        # print(f"[ORDINAL VALUES]: {[ord(c) for c in raw_line]}")


        if not line:
            if current_list_item_buffer:
                parsed_data.append({
                    'level1': current_level1,
                    'level2': current_level2,
                    'level3': current_level3,
                    'item_no': current_list_item_buffer['number'], # 'number' could be actual number or bullet symbol
                    'item_desc': current_list_item_buffer['description'].strip()
                })
                current_list_item_buffer = None
            # print("Skipping empty line.")
            continue

        match_level1 = pattern_level1.match(line)
        match_level2 = pattern_level2.match(line)
        match_level3 = pattern_level3.match(line)
        match_list_item = pattern_list_item.match(line)
        # Use raw_line for bullet item matching to preserve leading whitespace
        match_bullet_item = pattern_bullet_item.match(raw_line) 

        # More detailed match output (removed for cleaner console during GUI operations)
        # if match_level1: print(f"MATCH: Level 1 - '{match_level1.group(1)}'")
        # elif match_level2: print(f"MATCH: Level 2 - '{match_level2.group(1)}'")
        # elif match_level3: print(f"MATCH: Level 3 - '{match_level3.group(1)}'")
        # elif match_list_item: print(f"MATCH: Numbered List Item - No:{match_list_item.group(1)}, Desc:'{match_list_item.group(2)}'")
        # elif match_bullet_item: print(f"MATCH: Bullet List Item - Desc:'{match_bullet_item.group(1)}'")
        # else: print("NO MATCH for any known pattern.")


        # Finalize any buffered list item before processing a new heading or list item
        if current_list_item_buffer and (match_level1 or match_level2 or match_level3 or match_list_item or match_bullet_item):
            parsed_data.append({
                'level1': current_level1,
                'level2': current_level2,
                'level3': current_level3,
                'item_no': current_list_item_buffer['number'],
                'item_desc': current_list_item_buffer['description'].strip()
            })
            # print(f"Finalized buffered list item: {current_list_item_buffer['number']}")
            current_list_item_buffer = None

        if match_level1:
            current_level1 = match_level1.group(1)
            current_level2 = ""
            current_level3 = ""
            parsed_data.append({
                'level1': current_level1, 'level2': '', 'level3': '',
                'item_no': '', 'item_desc': ''
            })
            # print(f"Set L1: '{current_level1}'")
        elif match_level2:
            current_level2 = match_level2.group(1)
            current_level3 = ""
            parsed_data.append({
                'level1': current_level1, 'level2': current_level2, 'level3': '',
                'item_no': '', 'item_desc': ''
            })
            # print(f"Set L2: '{current_level2}' (under L1: '{current_level1}')")
        elif match_level3:
            current_level3 = match_level3.group(1)
            parsed_data.append({
                'level1': current_level1, 'level2': current_level2, 'level3': current_level3,
                'item_no': '', 'item_desc': ''
            })
            # print(f"Set L3: '{current_level3}' (under L1: '{current_level1}', L2: '{current_level2}')")
        elif match_list_item:
            # If a numbered list item starts, finalize any previous buffer
            if current_list_item_buffer:
                parsed_data.append({
                    'level1': current_level1,
                    'level2': current_level2,
                    'level3': current_level3,
                    'item_no': current_list_item_buffer['number'],
                    'item_desc': current_list_item_buffer['description'].strip()
                })
                # print(f"Finalized previous buffered item before new numbered list: {current_list_item_buffer['number']}")
            current_list_item_buffer = {
                'number': match_list_item.group(1), # This will be the actual number
                'description': match_list_item.group(2)
            }
            # print(f"Buffered numbered list item: No.{current_list_item_buffer['number']} Desc:'{current_list_item_buffer['description']}'")
        elif match_bullet_item: # NEW LOGIC FOR BULLET POINTS
            # If a bullet point starts, finalize any previous buffer
            if current_list_item_buffer:
                parsed_data.append({
                    'level1': current_level1,
                    'level2': current_level2,
                    'level3': current_level3,
                    'item_no': current_list_item_buffer['number'],
                    'item_desc': current_list_item_buffer['description'].strip()
                })
                # print(f"Finalized previous buffered item before new bullet list.")
            # For bullet points, we will use '•' as the item_no in the Excel
            current_list_item_buffer = {
                'number': '•', # Special symbol for bullet points
                'description': match_bullet_item.group(1)
            }
            # print(f"Buffered bullet list item: Desc:'{current_list_item_buffer['description']}'")
        else:
            if current_list_item_buffer:
                current_list_item_buffer['description'] += " " + line
                # print(f"Continued list item: '{current_list_item_buffer['description']}'")
            # If there's no match and no buffer, the line is effectively skipped for parsing.


    # Handle any remaining buffered list item after loop
    if current_list_item_buffer:
        parsed_data.append({
            'level1': current_level1,
            'level2': current_level2,
            'level3': current_level3,
            'item_no': current_list_item_buffer['number'],
            'item_desc': current_list_item_buffer['description'].strip()
        })
        # print(f"Finalized remaining buffered list item: {current_list_item_buffer['number']}")

    print(f"Parsed {len(parsed_data)} data rows")

    # Write data to sheet
    for row_data in parsed_data:
        sheet.append([
            row_data['level1'],
            row_data['level2'],
            row_data['level3'],
            row_data['item_no'],
            row_data['item_desc']
        ])

    # Merging function with improved error handling
    def merge_consecutive_cells(column_index, sheet):
        """Merge consecutive cells with the same content in a column"""
        if sheet.max_row < 2:
            return
            
        current_value = None
        start_row = None
        
        for row_num in range(2, sheet.max_row + 1):
            try:
                cell_value = sheet.cell(row=row_num, column=column_index).value
                
                if cell_value and str(cell_value).strip():
                    if current_value != cell_value:
                        # End previous merge if exists
                        if start_row and start_row < row_num - 1:
                            merge_range = f"{get_column_letter(column_index)}{start_row}:{get_column_letter(column_index)}{row_num-1}"
                            try:
                                sheet.merge_cells(merge_range)
                                merged_cell = sheet.cell(row=start_row, column=column_index)
                                merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                # print(f"Merged {merge_range}")
                            except Exception as e:
                                print(f"Error merging {merge_range}: {e}")
                            
                        current_value = cell_value
                        start_row = row_num
                else:
                    if start_row and start_row < row_num - 1:
                        merge_range = f"{get_column_letter(column_index)}{start_row}:{get_column_letter(column_index)}{row_num-1}"
                        try:
                            sheet.merge_cells(merge_range)
                            merged_cell = sheet.cell(row=start_row, column=column_index)
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            # print(f"Merged {merge_range}")
                        except Exception as e:
                            print(f"Error merging {merge_range}: {e}")
                    
                    current_value = None
                    start_row = None
            except Exception as e:
                print(f"Error processing row {row_num}, column {column_index}: {e}")
                continue
        
        # Handle final merge
        if start_row and start_row < sheet.max_row:
            merge_range = f"{get_column_letter(column_index)}{start_row}:{get_column_letter(column_index)}{sheet.max_row}"
            try:
                sheet.merge_cells(merge_range)
                merged_cell = sheet.cell(row=start_row, column=column_index)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # print(f"Final merge {merge_range}")
            except Exception as e:
                print(f"Error in final merge {merge_range}: {e}")

    # Perform merging with error handling
    print("Starting merging process...")
    try:
        merge_consecutive_cells(1, sheet)  # Level 1
        merge_consecutive_cells(2, sheet)  # Level 2  
        merge_consecutive_cells(3, sheet)  # Level 3
    except Exception as e:
        print(f"Error during merging: {e}")
        print("Continuing without merging...")

    # Set column widths
    column_widths = [25, 30, 35, 8, 60]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Set row heights
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 25

    # Apply formatting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, 6):
            try:
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                if row == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif col <= 3:  # Heading columns
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Content columns
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            except Exception as e:
                print(f"Error formatting cell {row},{col}: {e}")
                continue
    
    # Save the workbook to an in-memory BytesIO object instead of a file
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0) # Rewind the buffer to the beginning
    return excel_buffer

if __name__ == "__main__":
    # Start the Eel application
    # The 'index.html' file should be in a 'web' subfolder
    eel.start('index.html', size=(900, 700))
